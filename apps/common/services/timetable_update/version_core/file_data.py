import hashlib
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote

import requests
from openpyxl import load_workbook

from apps.common.models import Resource, FileVersion, Tag
from .stringlistanalyzer import StringListAnalyzer

logger = logging.getLogger(__name__)


class FileData:
    """
    Хранит все параметры файла расписания, извлечённые из пути и URL со страницы сайта.
    Предоставляет методы для создания записей Resource и FileVersion для БД.
    """

    _CONFIDENCE_VALUE = 0.8

    _DEGREE_WORDS = ["бакалавриат", "специалитет", "магистратура", "аспирантура", "степень"]
    _EDUCATION_FORM_WORDS = ["форма", "очная", "очно-заочная", "заочная"]
    _FACULTY_WORDS = [
        "факультет", "автоматизированных", "систем", "транспорта", "вооружений",
        "автомобильного", "технологии", "конструкционных", "материалов", "пищевых",
        "производств", "экономика", "управление", "электроника", "вычислительная",
        "техника", "xимико-технологический", "иностранный", "вечерний",
        "технологический", "инженерный", "кадры",
    ]
    _WORDS_TO_DELETE = ["автосохраненный", "копия"]
    _COURSE_WORDS = ["курс", "год"]
    _SENTENCE_DELIMITERS = ["_", " ", "(", ")", ",", ".", '"']
    _EXCEL_EXTENSION = [".xls", ".xlsx", ".xlsm"]

    def __init__(self, path: str, url: str, last_update: str) -> None:
        self.__path = path
        self.__url = url
        self.__last_changed = last_update
        self.__calc()

    def get_path(self) -> str:
        return self.__path

    def get_url(self) -> str:
        return self.__url

    def get_last_changed(self) -> str:
        return self.__last_changed

    def get_name(self) -> str:
        return self.__correct_name_from_path

    def get_mimetype(self) -> str:
        return self.__mimetype_from_url

    def get_file_name(self) -> str:
        """Возвращает краткое имя файла для сохранения на диск."""
        name = self.__name_from_url_with_mimetype
        words = name.split()
        abbr = "".join(word[0].upper() for word in words)
        suffix = Path(name).suffix
        return abbr + suffix

    def get_correct_path(self) -> str:
        path = self.__correct_path
        path_parts = path.strip("/").split("/")
        abbreviations = []
        for part in path_parts:
            if part.startswith("Курс"):
                abbr = "К" + part.split()[-1]
            else:
                words = [w for w in part.split() if len(w) > 2]
                abbr = "".join(w[0].upper() for w in words)
            abbreviations.append(abbr)
        return "/".join(abbreviations)

    def get_resource(self, type_timetable: str) -> Resource:
        """Создаёт и возвращает несохранённый объект Resource для БД."""
        resource = Resource()
        resource.path = self.get_correct_path()
        resource.name = self.get_name()
        resource.add_tags(*self._get_tags(type_timetable))
        resource.deprecated = False
        return resource

    def get_file_version(self, file_path: Path | str) -> FileVersion:
        """
        Создаёт и возвращает несохранённый объект FileVersion с хэшом содержимого файла.
        :param file_path: путь к скачанному локальному файлу
        """
        file_path = Path(file_path)
        if not file_path.is_file():
            raise FileNotFoundError(f"File not found: {file_path}")

        file_version = FileVersion()
        file_version.mimetype = file_path.suffix
        file_version.url = self.__url

        try:
            file_version.last_changed = datetime.strptime(self.__last_changed, "%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            file_version.last_changed = datetime.now()

        file_version.hashsum = self.__get_file_hash(file_path)
        return file_version

    # ------------------- ПРИВАТНЫЕ МЕТОДЫ ------------------- #

    def __calc(self) -> None:
        self.__name_from_path = self.get_file_name_from_path(self.__path)
        self.__mimetype_from_path = self.__get_mimetype(self.__path)
        self.__correct_name_from_path = self.get_correct_file_name(self.__name_from_path)

        self.__name_from_url_with_mimetype = self.get_file_name_from_path(self.__url, dell_mimetype=False)
        self.__name_from_url = self.get_file_name_from_path(self.__url)
        self.__mimetype_from_url = self.__get_mimetype(self.__url)
        self.__correct_name_from_url = self.get_correct_file_name(self.__name_from_url)

        self.__degree = self._get_degree(self.__path)
        self.__education_form = self._get_education_form(self.__path)
        self.__faculty = self._get_faculty(self.__path)
        self.__course = self._get_course_list(self.__name_from_path)
        self.__correct_path = self.__calc_correct_path(self.__path)

    def _get_json(self, type_timetable: str) -> str:
        return json.dumps({
            "type_timetable": type_timetable,
            "degree": self.__degree,
            "education_form": self.__education_form,
            "faculty": self.__faculty,
            "course": self.__course,
        }, indent=4, ensure_ascii=False)

    def _get_tags(self, type_timetable: str) -> list[Tag]:
        tags = [Tag(name=type_timetable, category="type_timetable")]

        for value, category in [
            (self.__degree, "degree"),
            (self.__education_form, "education_form"),
            (self.__faculty, "faculty"),
        ]:
            tags.append(Tag(name=value or "Неопределено", category=category))

        for course in (self.__course or ["Неопределено"]):
            tags.append(Tag(name=str(course), category="course"))

        return tags

    def __calc_correct_path(self, path: str, number_of_first_directories: int = 0) -> str:
        dirs = path.split("/")
        schedule_type = "Расписание занятий" if "занятий" in path else "Расписание экзаменов"
        new_path = dirs[:number_of_first_directories]
        new_path.append(schedule_type)
        new_path.append(self.__degree)
        new_path.append(self.__faculty)
        new_path.append(self.__education_form)
        new_path.append(self.__get_course_string(self.__course))
        return self.elements_to_path(new_path)

    @classmethod
    def __get_file_hash(cls, file_path: Path) -> str:
        if file_path.suffix in cls._EXCEL_EXTENSION:
            return cls.__get_excel_file_hash(file_path)
        return cls.__get_bin_file_hash(file_path)

    @classmethod
    def __get_excel_file_hash(cls, file_path: Path) -> str:
        if file_path.suffix.lower() == '.xls':
            import xlrd
            wb = xlrd.open_workbook(str(file_path))
            data = []
            for sheet in wb.sheets():
                for row in range(sheet.nrows):
                    data.append(tuple(sheet.row_values(row)))
            return hashlib.sha256(str(data).encode("utf-8")).hexdigest()
        else:
            wb = load_workbook(str(file_path), data_only=True)
            data = [tuple(row) for sheet in wb.worksheets for row in sheet.iter_rows(values_only=True)]
            return hashlib.sha256(str(data).encode("utf-8")).hexdigest()

    @staticmethod
    def __get_bin_file_hash(file_path: Path) -> str:
        sha256 = hashlib.sha256()
        with file_path.open("rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256.update(chunk)
        return sha256.hexdigest()

    def download_file(self, directory: Path | str, chunk_size: int = 8192) -> Path:
        """Скачивает файл по URL и сохраняет в указанную директорию."""
        response = requests.get(self.__url, timeout=30)
        if response.status_code != 200:
            raise Exception(f"File download error. Status: {response.status_code}, URL: {self.__url}")

        directory = Path(directory)
        directory.mkdir(parents=True, exist_ok=True)
        file_path = directory / self.get_file_name()

        with file_path.open("wb") as f:
            for chunk in response.iter_content(chunk_size=chunk_size):
                f.write(chunk)

        return file_path

    # ------------------- КЛАССОВЫЕ ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ------------------- #

    @classmethod
    def get_file_name_from_path(cls, path: str, dell_mimetype: bool = True) -> str:
        file_name = unquote(path.split("/")[-1])
        if dell_mimetype:
            file_name = re.sub(r"\.[А-ЯЁA-Zа-яёa-z]*$", "", file_name)
        return file_name

    @classmethod
    def get_correct_file_name(cls, file_name: str) -> str:
        name = cls.__remove_extra_spaces(file_name)
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(name), cls._WORDS_TO_DELETE)
        for word in analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1):
            name = name.replace(word, "")
        while True:
            before = name
            name = re.sub(r"-\s*$", "", name)
            name = re.sub(r"\(\s*\)", "", name)
            name = cls.__remove_extra_spaces(name)
            if name == before:
                break
        return name

    @classmethod
    def split_string_by_delimiters(cls, string: str, delimiters: list | None = None) -> list[str]:
        if delimiters is None:
            delimiters = cls._SENTENCE_DELIMITERS
        return re.split("|".join(map(re.escape, delimiters)), string)

    @classmethod
    def elements_to_path(cls, elements: list[str], base_path: str = "", is_file: bool = False) -> str:
        result = base_path
        for i, el in enumerate(elements):
            is_last = i + 1 >= len(elements)
            if el:
                result += el
                if not (is_last and is_file):
                    result += "/"
        return result

    @staticmethod
    def __remove_extra_spaces(string: str) -> str:
        return re.sub(r"\s+", " ", string).strip()

    @staticmethod
    def __get_mimetype(name: str) -> str:
        return name.split(".")[-1]

    @classmethod
    def _get_degree(cls, path: str | list) -> str:
        if isinstance(path, str):
            path = path.split("/")
        return cls.__get_best_element(path, cls.__get_degree_word_count)

    @classmethod
    def _get_education_form(cls, path: str | list) -> str:
        if isinstance(path, str):
            path = path.split("/")
        return cls.__get_best_element(path, cls.__get_education_form_word_count)

    @classmethod
    def _get_faculty(cls, path: str | list) -> str:
        if isinstance(path, str):
            path = path.split("/")
        return cls.__get_best_element(path, cls.__get_faculty_word_count)

    @classmethod
    def _get_course_list(cls, name: str) -> list[int]:
        """Извлекает список номеров курсов из имени файла."""
        parts = cls.split_string_by_delimiters(name.lower())
        result = []
        for i, part in enumerate(parts):
            if any(cw in part for cw in cls._COURSE_WORDS) and i + 1 < len(parts):
                try:
                    result.extend(cls.__parse_course_string(parts[i + 1]))
                except Exception:
                    pass
        return sorted(set(result))

    @staticmethod
    def __parse_course_string(string: str) -> list[int]:
        numbers = []
        for part in string.split(","):
            part = part.strip()
            if "-" in part:
                bounds = part.split("-")
                numbers.extend(range(int(bounds[0]), int(bounds[-1]) + 1))
            elif part.isdigit():
                numbers.append(int(part))
        return numbers

    @staticmethod
    def __get_course_string(courses: list[int]) -> str:
        if not courses:
            return "Неопределено"
        if len(courses) == 1:
            return f"Курс {courses[0]}"
        return f"Курс {courses[0]}-{courses[-1]}"

    @classmethod
    def __get_best_element(cls, elements: list[str], score_fn) -> str:
        best, best_score = "", 0
        for el in elements:
            score = score_fn(el)
            if score > best_score:
                best, best_score = el, score
        return best

    @classmethod
    def __get_degree_word_count(cls, string: str) -> int:
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(string.lower()), cls._DEGREE_WORDS)
        return len(analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1))

    @classmethod
    def __get_education_form_word_count(cls, string: str) -> int:
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(string.lower()), cls._EDUCATION_FORM_WORDS)
        return len(analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1))

    @classmethod
    def __get_faculty_word_count(cls, string: str) -> int:
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(string.lower()), cls._FACULTY_WORDS)
        return len(analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1))
